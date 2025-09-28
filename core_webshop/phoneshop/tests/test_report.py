from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth.models import User
from phoneshop.models import Shops, Products, Users, Workers, Sales
from django.utils import timezone
from datetime import date
from openpyxl import load_workbook
from io import BytesIO


class ExportReportExcelTest(TestCase):
    def setUp(self):
        self.client = Client()

        self.shop = Shops.objects.create(name="Test Shop")


        self.superuser = User.objects.create_superuser(
            username="admin", email="admin@admin.com", password="admin"
        )


        self.worker = Workers.objects.create(
            first_name="Admin",
            last_name="User",
            shop=self.shop,
            address="Some Address",
            birth_date="2000-01-01",
            phone_number="123456789",
            position="Manager",
            admin=True
        )


        self.user_phoneshop = Users.objects.create(user=self.superuser, worker=self.worker)

        self.url = reverse("export_report_excel")

    def login_and_set_session(self):
        self.client.login(username="admin", password="admin")
        session = self.client.session
        session['shop'] = self.shop.name
        session.save()

    def test_export_report_excel_response(self):
        self.login_and_set_session()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'], 
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def test_export_report_excel_filename(self):
        self.login_and_set_session()
        response = self.client.get(self.url)
        today = date.today()
        self.assertIn(f"report_{self.shop.name}_{today}.xlsx", response['Content-Disposition'])

    def test_export_report_excel_headers(self):
        self.login_and_set_session()
        response = self.client.get(self.url)
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ["Termék neve", "Szín", "Tárhely", "Mennyiség", "Ár"])

    def test_export_report_excel_includes_shop_sales(self):
        self.login_and_set_session()

        product = Products.objects.create(
            name="Iphone Test", category="Telefon", price=100000, colors=["Black"]
        )
        sale = Sales.objects.create(
            product=product,
            shop=self.shop,
            quantity=2,
            price=200000,
            color="Black",
            storage=64,
            selling_time=timezone.now(),
            costumer_name="Teszt Vevő",
            zip_code="1234",
            address="Valahol",
            city="Budapest",
            tax_number="12345678",
            user=self.user_phoneshop
        )

        response = self.client.get(self.url)
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        values = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
        self.assertIn([
            sale.product.name,
            sale.color,
            "64GB",
            "2db",
            "200000Ft"
        ], values)
